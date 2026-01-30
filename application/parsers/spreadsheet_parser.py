"""Spreadsheet parser module for parsing Excel files and creating Item objects"""

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl import load_workbook

from application.translations import TRANSLATIONS


@dataclass(frozen=True)
class SourceFromTLIPath:
    """Represents a row from order_path_properties table."""

    order_path_properties_id: Optional[int] = None
    order_path: str = ""
    order_change_path: str = ""  # legacy, no longer stored in DB
    java_code_wrapper: str = ""  # legacy, no longer stored in DB
    xtl_part_to_replace_850: str = ""
    xtl_part_to_paste_850: str = ""
    xtl_part_to_replace_860: str = ""
    xtl_part_to_paste_860: str = ""


@dataclass
class SourcingGroup:
    """Represents a row from sourcing_group_properties table, with a link to SourceFromTLIPath."""

    sourcing_group_properties_id: Optional[int] = None
    populate_method_name: str = ""
    map_name: str = ""
    order_path_properties_id: Optional[int] = None
    call_method_java_code: str = ""
    source_from_tli_path: Optional[SourceFromTLIPath] = None


@dataclass
class Item:
    """Item class representing parsed spreadsheet data"""
    
    # Spreadsheet-related properties
    spreadsheet_label: str = ""
    spreadsheet_edi_info_text: str = ""
    spreadsheet_edi_info_text_cleared: str = ""
    spreadsheet_usage: str = ""
    spreadsheet_min_max_text: str = ""
    spreadsheet_min: Optional[int] = None
    spreadsheet_max: Optional[int] = None
    spreadsheet_description: str = ""
    
    # EDI-related properties
    edi_segment: str = ""
    edi_element_number: str = ""
    edi_qualifier: str = ""
    
    # Item properties from DB
    item_properties_id: Optional[int] = None
    tli_value: str = ""
    rsx_tag_850: str = ""
    tli_tag_850: str = ""
    extra_record_defining_rsx_tag: str = ""
    extra_record_defining_qual: str = ""
    is_on_detail_level: bool = False
    is_partnumber: bool = False
    rsx_path_855: str = ""
    rsx_path_856: str = ""
    rsx_path_810: str = ""
    put_in_855: bool = False
    put_in_856: bool = False
    put_in_810: bool = False

    # References to sourcing group and order path
    order_path_properties_id: Optional[int] = None
    sourcing_group: Optional[SourcingGroup] = None
    
    # Parsing errors
    parsing_errors: List[str] = field(default_factory=list)
    
    @staticmethod
    def clear_edi_info(line: str) -> str:
        """
        Clear EDI info from line
        
        Args:
            line: Input line with EDI info
            
        Returns:
            Cleared EDI info string
        """
        line = line.strip()
        
        # If there is no colon, this is a simple value
        if ":" not in line:
            return line
        
        # Find all keys and their values
        # Key = non-whitespace characters followed by ':' and spaces
        # Value = everything until the next key (pattern: number followed by ':') or end of line
        # Also handle cases like "850: PO103 / 860: POC05" where we need to extract "PO103" (before "/")
        pairs = re.findall(
            r'(\S+):\s*([^:]+?)(?=\s*\d+:|$)',
            line
        )
        
        for key, value in pairs:
            if "850" in key:
                # Remove trailing "/" and whitespace if present
                value = value.strip()
                if value.endswith("/"):
                    value = value[:-1].strip()
                return value
        
        # If key 850 is not found, return the original line
        return line
    
    @staticmethod
    def normalize_segment(seg: str) -> str:
        """
                Normalize segment name.

                - Converts all variants that start with "P0" to "PO"
                    (so P01, P04, etc. become PO1, PO4)
                - Additional normalization rules can be added later if needed
        """
        if seg.startswith("P0"):
            return "PO" + seg[2:]
        return seg
    
    @staticmethod
    def parse_edi_info(text: str) -> Tuple[str, str, str]:
        """
        Parse EDI info text
        
        Args:
            text: EDI info text to parse
            
        Returns:
            Tuple of (edi_segment, edi_element_number, edi_qualifier)
        """
        text = text.strip()
        
        if not text:
            return "", "", ""

        # Format: "850: P0107/PO109 (IB)" or "P0107/PO109 (IB)" (after clear_edi_info)
        # Extract prefix (optional, e.g., "850:"), segment variants (e.g., "P0107/PO109"), and qualifier (e.g., "(IB)")
        m = re.match(r'^(?:\d+:\s*)?([^/]+)(?:/[^)]+)?\s*\(([A-Za-z0-9]+)\)$', text)
        if m:
            segment_part, qual = m.groups()
            # Parse segment part (e.g., "P0107" -> PO1, 07)
            seg_part = segment_part.strip()
            # Handle P0107 format
            if seg_part.startswith("P0") and len(seg_part) >= 4:
                # P0107 -> PO1, 07
                seg_digit = seg_part[2]  # Get digit after "P0" (index 2)
                el = seg_part[3:].zfill(2)  # Get remaining digits as element number (from index 3)
                seg = f"PO{seg_digit}"
                return seg, el, qual
            # Handle PO109 format (if segment_part is already PO1)
            elif seg_part.startswith("PO") and len(seg_part) >= 4:
                # PO109 -> PO1, 09
                seg = seg_part[:3]  # PO1
                el = seg_part[3:].zfill(2)  # 09
                return seg, el, qual

        # Special handling for values like P0401, P0402, P0101, etc.
        # P0401 -> seg = PO4, el = 01; P0101 -> PO1, 01; P0402 -> PO4, 02, etc.
        m = re.match(r'^P0(\d)(\d{2})$', text)
        if m:
            seg_digit, el = m.groups()
            seg = f"PO{seg_digit}"
            return seg, el, ""
        
        # Format: SEGNN (SEGMM = QUAL)
        m = re.match(r'^(\S+?)(\d+)\s*\(\s*(\S+?)(\d+)\s*=\s*([A-Za-z0-9]+)\s*\)$', text)
        if m:
            seg_part, digits, qseg, qel, qual = m.groups()
            # Special generic case for values like P0401, P0101, etc.
            # P0401 (N1=VN) -> seg = PO4, el = 01; P0109 (UP) -> PO1, 09, etc.
            if seg_part == "P" and len(digits) >= 3 and digits[0] == "0":
                seg = "PO" + digits[1]
                el = digits[-2:].zfill(2)
            # If the element number has 3 or more digits, the first digit belongs to the segment
            elif len(digits) >= 3:
                # Take the first digit for the segment and the last 2 for the element number
                # For example, N403 -> seg = N4, el = 03
                seg = seg_part + digits[0]
                el = digits[-2:]  # Last 2 digits
            elif len(digits) == 2:
                # If there are 2 digits, check the segment length
                # If the segment is longer than 1 character (e.g. PID, N1, N2), the segment is complete
                #   and both digits are the element number
                # If the segment is 1 character (e.g. N), the first digit goes to the segment, the second to the element number
                if len(seg_part) > 1:
                    # Segment is complete, both digits are the element number
                    # For example, PID05 (PID02=08) -> seg = PID, el = 05
                    seg = seg_part
                    el = digits.zfill(2)
                else:
                    # Segment is 1 character, the first digit goes to the segment
                    # For example, N45 (N402=08) -> seg = N4, el = 05
                    seg = seg_part + digits[0]
                    el = digits[1].zfill(2)
            else:
                # If there is 1 digit, it goes to the element number
                seg = seg_part
                el = digits.zfill(2)
            seg = Item.normalize_segment(seg)
            return seg, el, qual
        
        # Format: SEGNN (QUAL)
        m = re.match(r'^(\S+?)(\d+)\s*\(\s*([A-Za-z0-9]+)\s*\)$', text)
        if m:
            seg_part, digits, qual = m.groups()
            # Special generic case for values like P0401, P0101, P0109, etc.
            # P0401 (UP) -> seg = PO4, el = 01; P0109 (UP) -> PO1, 09, etc.
            if seg_part == "P" and len(digits) >= 3 and digits[0] == "0":
                seg = "PO" + digits[1]
                el = digits[-2:].zfill(2)
            # If the element number has 3 or more digits, the first digit belongs to the segment
            elif len(digits) >= 3:
                # Take the first digit for the segment and the last 2 for the element number
                # For example, N403 -> seg = N4, el = 03
                seg = seg_part + digits[0]
                el = digits[-2:]  # Last 2 digits
            elif len(digits) == 2:
                # If there are 2 digits, check the segment length
                # If the segment is longer than 1 character (e.g. PID, N1, N2), the segment is complete
                #   and both digits are the element number
                # If the segment is 1 character (e.g. N), the first digit goes to the segment, the second to the element number
                if len(seg_part) > 1:
                    # Segment is complete, both digits are the element number
                    # For example, PID05 (08) -> seg = PID, el = 05
                    seg = seg_part
                    el = digits.zfill(2)
                else:
                    # Segment is 1 character, the first digit goes to the segment
                    # For example, N45 (08) -> seg = N4, el = 05
                    seg = seg_part + digits[0]
                    el = digits[1].zfill(2)
            else:
                # If there is 1 digit, it goes to the element number
                seg = seg_part
                el = digits.zfill(2)
            seg = Item.normalize_segment(seg)
            return seg, el, qual
        
        # Format: SEGNN (for example N404 -> N4, 04)
        # First try to find a segment that already ends with a digit (for example N4, PO1)
        # If the element number has more than 2 digits, take the last 2
        # Important: this pattern should trigger only if the element number has 2 or more digits
        # to avoid parsing "PID05" incorrectly as "PID0" + "5"
        m = re.match(r'^([A-Za-z]+\d)(\d{2,})$', text)
        if m:
            seg, el = m.groups()
            seg = Item.normalize_segment(seg)
            # Limit edi_element_number to 2 digits (take the last 2)
            # For example, N404 -> N4, 04 (last 2 digits from 404)
            el = el[-2:].zfill(2) if len(el) > 2 else el.zfill(2)
            return seg, el, ""
        
        # Format: SEGNN (when the segment has no trailing digit, for example N404 where N is the segment and 404 is the number)
        # But if the number has 3 or more digits, it may actually be N4 + 04
        m = re.match(r'^([A-Za-z]+)(\d+)$', text)
        if m:
            seg_part, digits = m.groups()
            # Special generic case for values like P0401, P0101, etc.
            # P0401 -> seg = PO4, el = 01; P0101 -> PO1, 01; P0402 -> PO4, 02, etc.
            if seg_part == "P" and len(digits) >= 3 and digits[0] == "0":
                seg = "PO" + digits[1]
                el = digits[-2:].zfill(2)
            # General logic for other segments
            elif len(digits) >= 3:
                # Take the first digit for the segment and the last 2 for the element number
                # For example, N404 -> seg = N4, el = 04
                seg = seg_part + digits[0]
                el = digits[-2:]  # Last 2 digits
            elif len(digits) == 2:
                # If there are 2 digits, check the segment length
                # If the segment is longer than 1 character (e.g. PID, N1, N2), the segment is complete
                #   and both digits are the element number
                # If the segment is 1 character (e.g. N), the first digit goes to the segment, the second to the element number
                if len(seg_part) > 1:
                    # Segment is complete, both digits are the element number
                    # For example, PID05 -> seg = PID, el = 05
                    seg = seg_part
                    el = digits.zfill(2)
                else:
                    # Segment is 1 character, the first digit goes to the segment
                    # For example, N45 -> seg = N4, el = 05
                    seg = seg_part + digits[0]
                    el = digits[1].zfill(2)
            else:
                # If there is 1 digit, it goes to the element number but the segment stays unchanged
                # For example, N4 -> seg = N, el = 04
                seg = seg_part
                el = digits.zfill(2)
            seg = Item.normalize_segment(seg)
            return seg, el, ""
        
        # Nothing matched → return empty fields
        return "", "", ""


class SpreadsheetParser:
    """Parser for Excel spreadsheet files"""
    
    def __init__(self, database, language: str = "UA"):
        """
        Initialize parser
        
        Args:
            database: Database instance for matching items
            language: Language code ("UA" or "EN") for error messages
        """
        self.database = database
        self.language = language
        self.t = TRANSLATIONS.get(language, TRANSLATIONS["UA"])
    
    def parse(self, file_path: Path) -> Tuple[List[Item], bool, Optional[str]]:
        """
        Parse spreadsheet file and create Item objects
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Tuple of (items_list, success, error_message)
        """
        items: List[Item] = []
        all_errors: List[str] = []
        
        try:
            # Load workbook
            workbook = load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            # Get max column (starting from column B = 2)
            max_col = sheet.max_column
            
            # Iterate through columns starting from B (index 2)
            for col_idx in range(2, max_col + 1):
                item = Item()
                column_errors = []
                
                # Get values from rows 1-5
                row1_value = self._get_cell_value(sheet, 1, col_idx)
                row2_value = self._get_cell_value(sheet, 2, col_idx)
                row3_value = self._get_cell_value(sheet, 3, col_idx)
                row4_value = self._get_cell_value(sheet, 4, col_idx)
                row5_value = self._get_cell_value(sheet, 5, col_idx)
                
                # Row 1: spreadsheet_label (required)
                if not row1_value or str(row1_value).strip() == "":
                    column_errors.append(
                        f"{self.t['error_column']} {self._column_letter(col_idx)}: "
                        f"{self.t['error_empty_field'].format(field='spreadsheet_label')}"
                    )
                else:
                    item.spreadsheet_label = str(row1_value).strip()
                
                # Row 2: spreadsheet_edi_info_text (required)
                if not row2_value or str(row2_value).strip() == "":
                    column_errors.append(
                        f"{self.t['error_column']} {self._column_letter(col_idx)}: "
                        f"{self.t['error_empty_field'].format(field='spreadsheet_edi_info_text')}"
                    )
                else:
                    item.spreadsheet_edi_info_text = str(row2_value).strip()
                
                # Row 3: spreadsheet_usage (required)
                if not row3_value or str(row3_value).strip() == "":
                    column_errors.append(
                        f"{self.t['error_column']} {self._column_letter(col_idx)}: "
                        f"{self.t['error_empty_field'].format(field='spreadsheet_usage')}"
                    )
                else:
                    item.spreadsheet_usage = str(row3_value).strip()
                
                # Row 4: spreadsheet_min_max_text (optional)
                if row4_value and str(row4_value).strip():
                    item.spreadsheet_min_max_text = str(row4_value).strip()
                    # Parse min/max only if field is not empty
                    min_max_errors = self._parse_min_max(item, col_idx)
                    column_errors.extend(min_max_errors)
                else:
                    # If field is empty and spreadsheet_label ends with "UOM", set min=2, max=2
                    if item.spreadsheet_label and item.spreadsheet_label.strip().upper().endswith("UOM"):
                        item.spreadsheet_min = 2
                        item.spreadsheet_max = 2
                
                # Row 5: spreadsheet_description (optional)
                if row5_value:
                    item.spreadsheet_description = str(row5_value).strip()
                
                # If there are errors in required fields, skip further processing
                if column_errors:
                    item.parsing_errors = column_errors
                    all_errors.extend(column_errors)
                    items.append(item)
                    continue
                
                # Check if spreadsheet_edi_info_text is "Blank" (case-insensitive)
                is_blank = item.spreadsheet_edi_info_text.strip().lower() == "blank"
                
                if is_blank:
                    # Special handling for "Blank" values
                    # Don't parse EDI info or search in database
                    # Set tli_tag_850 from spreadsheet_label (letters only, first letter uppercase, rest keep original case)
                    if item.spreadsheet_label:
                        # Extract only letters from spreadsheet_label
                        letters_only = ''.join(c for c in item.spreadsheet_label if c.isalpha())
                        # Make first letter uppercase, rest keep original case
                        if letters_only:
                            item.tli_tag_850 = letters_only[0].upper() + letters_only[1:]
                        else:
                            item.tli_tag_850 = ""
                    else:
                        item.tli_tag_850 = ""
                    
                    # Set tli_value to empty string
                    item.tli_value = ""
                    
                    # Skip EDI parsing and database matching
                    item.parsing_errors = column_errors
                    items.append(item)
                    continue
                
                # Check if spreadsheet_label contains EDI info in format "850: ... / 860: ..."
                # If so, parse it and use it instead of spreadsheet_edi_info_text
                edi_info_from_label = None
                if item.spreadsheet_label and "850:" in item.spreadsheet_label:
                    try:
                        # Extract EDI info from label (e.g., "850: PO103 / 860: POC05" -> "PO103")
                        cleared_label = Item.clear_edi_info(item.spreadsheet_label)
                        if cleared_label and cleared_label != item.spreadsheet_label:
                            # Try to parse it
                            seg, el, qual = Item.parse_edi_info(cleared_label)
                            if seg and el:
                                # Successfully parsed from label, use this
                                edi_info_from_label = (seg, el, qual)
                    except Exception:
                        # If parsing from label fails, continue with normal flow
                        pass
                
                # Parse spreadsheet_edi_info_text_cleared
                try:
                    item.spreadsheet_edi_info_text_cleared = Item.clear_edi_info(item.spreadsheet_edi_info_text)
                except Exception as e:
                    error_msg = (
                        f"{self.t['error_column']} {self._column_letter(col_idx)}: "
                        f"{self.t['error_parse_spreadsheet_edi_info']}: {str(e)}"
                    )
                    column_errors.append(error_msg)
                    all_errors.append(error_msg)
                    item.parsing_errors = column_errors
                    items.append(item)
                    continue
                
                # Special handling for N104 with conditions in parentheses
                special_n104_handled = False
                cleared_text = item.spreadsheet_edi_info_text_cleared.strip()

                if cleared_text:
                    # Case 1: N104 (N101=VN and N103=92) -> segment N1, element 04, qualifier taken from N101
                    m = re.match(
                        r'^N104\s*\(\s*N101\s*=\s*([A-Za-z0-9]+)\s+and\s+N103\s*=\s*([A-Za-z0-9]+)\s*\)$',
                        cleared_text,
                        re.IGNORECASE,
                    )
                    if m:
                        qual_from_n101 = m.group(1).strip()
                        item.edi_segment = "N1"
                        item.edi_element_number = "04"
                        item.edi_qualifier = qual_from_n101
                        special_n104_handled = True
                    else:
                        # Case 2: N104 (N103=92) -> segment N1, element 04, qualifier is inherited
                        m = re.match(
                            r'^N104\s*\(\s*N103\s*=\s*([A-Za-z0-9]+)\s*\)$',
                            cleared_text,
                            re.IGNORECASE,
                        )
                        if m:
                            # Take the previous Item if it exists
                            if items and items[-1].edi_segment == "N1":
                                prev_item = items[-1]
                                item.edi_segment = "N1"
                                item.edi_element_number = "04"
                                item.edi_qualifier = prev_item.edi_qualifier
                                special_n104_handled = True
                            else:
                                # If there is no previous N1 Item, treat it as a parsing error
                                error_msg = (
                                    f"{self.t['error_column']} {self._column_letter(col_idx)}: "
                                    f"{self.t['error_n104_missing_previous_qualifier']}"
                                )
                                column_errors.append(error_msg)
                                all_errors.append(error_msg)
                                item.parsing_errors = column_errors
                                items.append(item)
                                continue

                # Parse EDI info (for all other cases)
                if not special_n104_handled:
                    # If EDI info was successfully parsed from spreadsheet_label, use it
                    if edi_info_from_label:
                        item.edi_segment, item.edi_element_number, item.edi_qualifier = edi_info_from_label
                    else:
                        # Otherwise, parse from spreadsheet_edi_info_text
                        try:
                            edi_segment, edi_element_number, edi_qualifier = Item.parse_edi_info(
                                item.spreadsheet_edi_info_text_cleared
                            )
                            item.edi_segment = edi_segment
                            item.edi_element_number = edi_element_number
                            item.edi_qualifier = edi_qualifier
                        except Exception as e:
                            error_msg = (
                                f"{self.t['error_column']} {self._column_letter(col_idx)}: "
                                f"{self.t['error_parse_edi_info']}: {str(e)}"
                            )
                            column_errors.append(error_msg)
                            all_errors.append(error_msg)
                            item.parsing_errors = column_errors
                            items.append(item)
                            continue
                    
                    # Check if we successfully parsed at least edi_segment and edi_element_number
                    if not item.edi_segment or not item.edi_element_number:
                        error_msg = (
                            f"{self.t['error_column']} {self._column_letter(col_idx)}: "
                            f"{self.t['error_failed_to_parse_edi_fields'].format(text=item.spreadsheet_edi_info_text_cleared)}"
                        )
                        column_errors.append(error_msg)
                        all_errors.append(error_msg)
                
                # Match with database
                # Only attempt matching if we have edi_segment and edi_element_number
                if item.edi_segment and item.edi_element_number:
                    match_errors = self._match_with_database(item, col_idx)
                    column_errors.extend(match_errors)
                    all_errors.extend(match_errors)
                
                item.parsing_errors = column_errors
                items.append(item)
            
            workbook.close()
            
            # Handle duplicate EDI combinations by adding sequential numbers to tli_tag_850
            self._handle_duplicate_edi_combinations(items)
            
            # Check if parsing was successful
            success = len(all_errors) == 0
            error_message = "\n".join(all_errors) if all_errors else None
            
            return items, success, error_message
            
        except Exception as e:
            return [], False, f"{self.t['error_read_file']}: {str(e)}"
    
    def _handle_duplicate_edi_combinations(self, items: List[Item]) -> None:
        """
        Handle duplicate EDI combinations by adding sequential numbers to tli_tag_850.
        
        For items with the same combination of edi_segment, edi_element_number, edi_qualifier,
        the second and subsequent items will have a sequential number (2, 3, 4, etc.) appended
        to their tli_tag_850.
        
        Args:
            items: List of parsed Item objects
        """
        # Dictionary to track occurrences of EDI combinations
        # Key: tuple of (edi_segment, edi_element_number, edi_qualifier)
        # Value: list of indices in items list where this combination appears
        edi_combination_indices: Dict[Tuple[str, str, str], List[int]] = {}
        
        # First pass: collect all items with EDI fields and group by combination
        for idx, item in enumerate(items):
            # Skip items without EDI fields (e.g., "Blank" cases or parsing errors)
            if not item.edi_segment or not item.edi_element_number:
                continue
            
            # Create key from EDI combination
            edi_key = (
                item.edi_segment,
                item.edi_element_number,
                item.edi_qualifier or ""  # Use empty string if qualifier is None
            )
            
            if edi_key not in edi_combination_indices:
                edi_combination_indices[edi_key] = []
            edi_combination_indices[edi_key].append(idx)
        
        # Second pass: add sequential numbers to duplicates
        for edi_key, indices in edi_combination_indices.items():
            # Only process if there are duplicates (more than one occurrence)
            if len(indices) > 1:
                # Sort indices to process in order
                indices.sort()
                
                # First occurrence keeps original tli_tag_850
                # Subsequent occurrences get sequential numbers appended
                for seq_num, idx in enumerate(indices[1:], start=2):
                    item = items[idx]
                    if item.tli_tag_850:
                        item.tli_tag_850 = f"{item.tli_tag_850}{seq_num}"
    
    def _get_cell_value(self, sheet, row: int, col: int) -> Optional[str]:
        """Get cell value as string"""
        cell = sheet.cell(row=row, column=col)
        if cell.value is None:
            return None
        return str(cell.value)
    
    def _column_letter(self, col_idx: int) -> str:
        """Convert column index to letter (1=A, 2=B, etc.)"""
        return openpyxl.utils.get_column_letter(col_idx)
    
    def _parse_min_max(self, item: Item, col_idx: int) -> List[str]:
        """Parse min/max from spreadsheet_min_max_text (optional field)"""
        errors = []
        text = item.spreadsheet_min_max_text.strip()
        
        # Format: "min=1, max=50" or similar
        min_match = re.search(r'min\s*=\s*(\d+)', text, re.IGNORECASE)
        max_match = re.search(r'max\s*=\s*(\d+)', text, re.IGNORECASE)
        
        # Check if text contains min= or max= keywords (field is optional, so only validate if keywords are present)
        has_min_keyword = bool(re.search(r'min\s*=', text, re.IGNORECASE))
        has_max_keyword = bool(re.search(r'max\s*=', text, re.IGNORECASE))
        
        if min_match:
            try:
                item.spreadsheet_min = int(min_match.group(1))
            except ValueError:
                errors.append(
                    f"{self.t['error_column']} {self._column_letter(col_idx)}: "
                    f"{self.t['error_invalid_min_format']} '{text}'"
                )
        elif has_min_keyword:
            # Only show error if min= keyword is present but value is invalid
            errors.append(
                f"{self.t['error_column']} {self._column_letter(col_idx)}: "
                f"{self.t['error_min_not_found']} '{text}'"
            )
        
        if max_match:
            try:
                item.spreadsheet_max = int(max_match.group(1))
            except ValueError:
                errors.append(
                    f"{self.t['error_column']} {self._column_letter(col_idx)}: "
                    f"{self.t['error_invalid_max_format']} '{text}'"
                )
        elif has_max_keyword:
            # Only show error if max= keyword is present but value is invalid
            errors.append(
                f"{self.t['error_column']} {self._column_letter(col_idx)}: "
                f"{self.t['error_max_not_found']} '{text}'"
            )
        
        return errors
    
    def _match_with_database(self, item: Item, col_idx: int) -> List[str]:
        """Match item with database records"""
        errors = []
        
        # Get all items from database
        db_items = self.database.get_all_items()
        
        # Normalize edi_element_number for matching
        edi_element_num = item.edi_element_number
        
        # Special case: if edi_segment=PO1 and edi_element_number>06, treat as 07
        if item.edi_segment == "PO1" and edi_element_num:
            try:
                if int(edi_element_num) > 6:
                    edi_element_num = "07"
            except ValueError:
                # If edi_element_number is not a valid number, keep original value
                pass
        
        # Find matches
        matches = []
        for db_item in db_items:
            db_segment = db_item.get("edi_segment", "")
            db_element_raw = db_item.get("edi_element_number", "")
            db_qualifier = db_item.get("edi_qualifier") or ""
            item_qualifier = item.edi_qualifier or ""
            
            # Normalize element numbers for comparison (convert both to int)
            # Database stores as INTEGER (e.g., 2), parser returns as string with leading zeros (e.g., "02")
            try:
                db_element_int = int(db_element_raw) if db_element_raw != "" else None
                item_element_int = int(edi_element_num) if edi_element_num else None
                element_match = db_element_int == item_element_int if (db_element_int is not None and item_element_int is not None) else False
            except (ValueError, TypeError):
                # If conversion fails, fall back to string comparison
                db_element = str(db_element_raw)
                element_match = db_element == edi_element_num
            
            if db_segment == item.edi_segment and element_match:
                # Match qualifier if both are present, or if both are empty
                if (db_qualifier and item_qualifier and db_qualifier == item_qualifier) or \
                   (not db_qualifier and not item_qualifier):
                    matches.append(db_item)
        
        if len(matches) == 0:
            errors.append(
                f"{self.t['error_column']} {self._column_letter(col_idx)}: "
                f"{self.t['error_no_match']} "
                f"edi_segment={item.edi_segment}, edi_element_number={item.edi_element_number}, "
                f"edi_qualifier={item.edi_qualifier or self.t['error_empty_qualifier']}"
            )
        elif len(matches) > 1:
            errors.append(
                f"{self.t['error_column']} {self._column_letter(col_idx)}: "
                f"{self.t['error_multiple_matches']} "
                f"edi_segment={item.edi_segment}, edi_element_number={item.edi_element_number}, "
                f"edi_qualifier={item.edi_qualifier or self.t['error_empty_qualifier']}"
            )
        else:
            # Single match - fill item properties
            match = matches[0]
            item.item_properties_id = match.get("item_properties_id")
            item.tli_value = match.get("TLI_value", "")
            item.rsx_tag_850 = match.get("850_RSX_tag", "")
            item.tli_tag_850 = match.get("850_TLI_tag", "")
            item.extra_record_defining_rsx_tag = match.get("extra_record_defining_rsx_tag") or ""
            item.extra_record_defining_qual = match.get("extra_record_defining_qual") or ""
            item.is_on_detail_level = bool(match.get("is_on_detail_level", False))
            item.is_partnumber = bool(match.get("is_partnumber", False))
            item.rsx_path_855 = match.get("855_RSX_path", "")
            item.rsx_path_856 = match.get("856_RSX_path", "")
            item.rsx_path_810 = match.get("810_RSX_path", "")
            item.put_in_855 = bool(match.get("put_in_855_by_default", False))
            item.put_in_856 = bool(match.get("put_in_856_by_default", False))
            item.put_in_810 = bool(match.get("put_in_810_by_default", False))

            # Get sourcing group and order path info
            sourcing_group_id = match.get("sourcing_group_properties_id")
            if sourcing_group_id:
                sg_row = self.database.get_sourcing_group(sourcing_group_id)
                if sg_row:
                    order_path_id = sg_row.get("order_path_properties_id")
                    source_path_obj: Optional[SourceFromTLIPath] = None

                    if order_path_id:
                        item.order_path_properties_id = order_path_id
                        op_row = self.database.get_order_path(order_path_id)
                        if op_row:
                            source_path_obj = SourceFromTLIPath(
                                order_path_properties_id=op_row.get("order_path_properties_id"),
                                order_path=op_row.get("order_path", ""),
                                order_change_path=op_row.get("order_change_path", ""),  # legacy
                                java_code_wrapper=op_row.get("java_code_wrapper", ""),  # legacy
                                xtl_part_to_replace_850=op_row.get("xtl_part_to_replace_850", ""),
                                xtl_part_to_paste_850=op_row.get("xtl_part_to_paste_850", ""),
                                xtl_part_to_replace_860=op_row.get("xtl_part_to_replace_860", ""),
                                xtl_part_to_paste_860=op_row.get("xtl_part_to_paste_860", ""),
                            )

                    item.sourcing_group = SourcingGroup(
                        sourcing_group_properties_id=sg_row.get("sourcing_group_properties_id"),
                        populate_method_name=sg_row.get("populate_method_name", ""),
                        map_name=sg_row.get("map_name", ""),
                        order_path_properties_id=sg_row.get("order_path_properties_id"),
                        call_method_java_code=sg_row.get("call_method_java_code", ""),
                        source_from_tli_path=source_path_obj,
                    )
        
        return errors

