"""Spreadsheet parser module for parsing Excel files and creating Item objects"""

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional, Tuple

import openpyxl
from openpyxl import load_workbook

from application.translations import TRANSLATIONS


@dataclass(frozen=True)
class SourceFromTLIPath:
    """Represents a row from order_path_properties table."""

    order_path_properties_id: Optional[int] = None
    order_path: str = ""
    order_change_path: str = ""  # legacy, no longer stored in DB
    java_code_wrapper: str = ""  # legacy, no longer stored in DB
    xtl_part_to_replace_850: str = ""
    xtl_part_to_paste_850: str = ""
    xtl_part_to_replace_860: str = ""
    xtl_part_to_paste_860: str = ""


@dataclass
class SourcingGroup:
    """Represents a row from sourcing_group_properties table, with a link to SourceFromTLIPath."""

    sourcing_group_properties_id: Optional[int] = None
    populate_method_name: str = ""
    map_name: str = ""
    order_path_properties_id: Optional[int] = None
    call_method_java_code: str = ""
    source_from_tli_path: Optional[SourceFromTLIPath] = None


@dataclass
class Item:
    """Item class representing parsed spreadsheet data"""
    
    # Spreadsheet-related properties
    spreadsheet_label: str = ""
    spreadsheet_edi_info_text: str = ""
    spreadsheet_edi_info_text_cleared: str = ""
    spreadsheet_usage: str = ""
    spreadsheet_min_max_text: str = ""
    spreadsheet_min: Optional[int] = None
    spreadsheet_max: Optional[int] = None
    spreadsheet_description: str = ""
    
    # EDI-related properties
    edi_segment: str = ""
    edi_element_number: str = ""
    edi_qualifier: str = ""
    
    # Item properties from DB
    item_properties_id: Optional[int] = None
    tli_value: str = ""
    rsx_tag_850: str = ""
    tli_tag_850: str = ""
    extra_record_defining_rsx_tag: str = ""
    extra_record_defining_qual: str = ""
    is_on_detail_level: bool = False
    is_partnumber: bool = False
    rsx_path_855: str = ""
    rsx_path_856: str = ""
    rsx_path_810: str = ""
    put_in_855: bool = False
    put_in_856: bool = False
    put_in_810: bool = False

    # References to sourcing group and order path
    order_path_properties_id: Optional[int] = None
    sourcing_group: Optional[SourcingGroup] = None
    
    # Parsing errors
    parsing_errors: List[str] = field(default_factory=list)
    
    @staticmethod
    def clear_edi_info(line: str) -> str:
        """
        Clear EDI info from line
        
        Args:
            line: Input line with EDI info
            
        Returns:
            Cleared EDI info string
        """
        line = line.strip()
        
        # If there is no colon, this is a simple value
        if ":" not in line:
            return line
        
        # Find all keys and their values
        # Key = non-whitespace characters followed by ':' and spaces
        # Value = everything until the next key or end of line
        pairs = re.findall(
            r'(\S+):\s*([^:]+?)(?=\s*\S+:|$)',
            line
        )
        
        for key, value in pairs:
            if "850" in key:
                return value.strip()
        
        # If key 850 is not found, return the original line
        return line
    
    @staticmethod
    def normalize_segment(seg: str) -> str:
        """
                Normalize segment name.

                - Converts all variants that start with "P0" to "PO"
                    (so P01, P04, etc. become PO1, PO4)
                - Additional normalization rules can be added later if needed
        """
        if seg.startswith("P0"):
            return "PO" + seg[2:]
        return seg
    
    @staticmethod
    def parse_edi_info(text: str) -> Tuple[str, str, str]:
        """
        Parse EDI info text
        
        Args:
            text: EDI info text to parse
            
        Returns:
            Tuple of (edi_segment, edi_element_number, edi_qualifier)
        """
        text = text.strip()
        
        if not text:
            return "", "", ""

        # Special handling for values like P0401, P0402, P0101, etc.
        # P0401 -> seg = PO4, el = 01; P0101 -> PO1, 01; P0402 -> PO4, 02, etc.
        m = re.match(r'^P0(\d)(\d{2})$', text)
        if m:
            seg_digit, el = m.groups()
            seg = f"PO{seg_digit}"
            return seg, el, ""
        
        # Format: SEGNN (SEGMM = QUAL)
        m = re.match(r'^(\S+?)(\d+)\s*\(\s*(\S+?)(\d+)\s*=\s*([A-Za-z0-9]+)\s*\)$', text)
        if m:
            seg_part, digits, qseg, qel, qual = m.groups()
            # If the element number has 3 or more digits, the first digit belongs to the segment
            if len(digits) >= 3:
                # Take the first digit for the segment and the last 2 for the element number
                # For example, N403 -> seg = N4, el = 03
                seg = seg_part + digits[0]
                el = digits[-2:]  # Last 2 digits
            elif len(digits) == 2:
                # If there are 2 digits, check the segment length
                # If the segment is longer than 1 character (e.g. PID, N1, N2), the segment is complete
                #   and both digits are the element number
                # If the segment is 1 character (e.g. N), the first digit goes to the segment, the second to the element number
                if len(seg_part) > 1:
                    # Segment is complete, both digits are the element number
                    # For example, PID05 (PID02=08) -> seg = PID, el = 05
                    seg = seg_part
                    el = digits.zfill(2)
                else:
                    # Segment is 1 character, the first digit goes to the segment
                    # For example, N45 (N402=08) -> seg = N4, el = 05
                    seg = seg_part + digits[0]
                    el = digits[1].zfill(2)
            else:
                # If there is 1 digit, it goes to the element number
                seg = seg_part
                el = digits.zfill(2)
            seg = Item.normalize_segment(seg)
            return seg, el, qual
        
        # Format: SEGNN (QUAL)
        m = re.match(r'^(\S+?)(\d+)\s*\(\s*([A-Za-z0-9]+)\s*\)$', text)
        if m:
            seg_part, digits, qual = m.groups()
            # If the element number has 3 or more digits, the first digit belongs to the segment
            if len(digits) >= 3:
                # Take the first digit for the segment and the last 2 for the element number
                # For example, N403 -> seg = N4, el = 03
                seg = seg_part + digits[0]
                el = digits[-2:]  # Last 2 digits
            elif len(digits) == 2:
                # If there are 2 digits, check the segment length
                # If the segment is longer than 1 character (e.g. PID, N1, N2), the segment is complete
                #   and both digits are the element number
                # If the segment is 1 character (e.g. N), the first digit goes to the segment, the second to the element number
                if len(seg_part) > 1:
                    # Segment is complete, both digits are the element number
                    # For example, PID05 (08) -> seg = PID, el = 05
                    seg = seg_part
                    el = digits.zfill(2)
                else:
                    # Segment is 1 character, the first digit goes to the segment
                    # For example, N45 (08) -> seg = N4, el = 05
                    seg = seg_part + digits[0]
                    el = digits[1].zfill(2)
            else:
                # If there is 1 digit, it goes to the element number
                seg = seg_part
                el = digits.zfill(2)
            seg = Item.normalize_segment(seg)
            return seg, el, qual
        
        # Format: SEGNN (for example N404 -> N4, 04)
        # First try to find a segment that already ends with a digit (for example N4, PO1)
        # If the element number has more than 2 digits, take the last 2
        # Important: this pattern should trigger only if the element number has 2 or more digits
        # to avoid parsing "PID05" incorrectly as "PID0" + "5"
        m = re.match(r'^([A-Za-z]+\d)(\d{2,})$', text)
        if m:
            seg, el = m.groups()
            seg = Item.normalize_segment(seg)
            # Limit edi_element_number to 2 digits (take the last 2)
            # For example, N404 -> N4, 04 (last 2 digits from 404)
            el = el[-2:].zfill(2) if len(el) > 2 else el.zfill(2)
            return seg, el, ""
        
        # Format: SEGNN (when the segment has no trailing digit, for example N404 where N is the segment and 404 is the number)
        # But if the number has 3 or more digits, it may actually be N4 + 04
        m = re.match(r'^([A-Za-z]+)(\d+)$', text)
        if m:
            seg_part, digits = m.groups()
            # Special generic case for values like P0401, P0101, etc.
            # P0401 -> seg = PO4, el = 01; P0101 -> PO1, 01; P0402 -> PO4, 02, etc.
            if seg_part == "P" and len(digits) >= 3 and digits[0] == "0":
                seg = "PO" + digits[1]
                el = digits[-2:].zfill(2)
            # General logic for other segments
            elif len(digits) >= 3:
                # Take the first digit for the segment and the last 2 for the element number
                # For example, N404 -> seg = N4, el = 04
                seg = seg_part + digits[0]
                el = digits[-2:]  # Last 2 digits
            elif len(digits) == 2:
                # If there are 2 digits, check the segment length
                # If the segment is longer than 1 character (e.g. PID, N1, N2), the segment is complete
                #   and both digits are the element number
                # If the segment is 1 character (e.g. N), the first digit goes to the segment, the second to the element number
                if len(seg_part) > 1:
                    # Segment is complete, both digits are the element number
                    # For example, PID05 -> seg = PID, el = 05
                    seg = seg_part
                    el = digits.zfill(2)
                else:
                    # Segment is 1 character, the first digit goes to the segment
                    # For example, N45 -> seg = N4, el = 05
                    seg = seg_part + digits[0]
                    el = digits[1].zfill(2)
            else:
                # If there is 1 digit, it goes to the element number but the segment stays unchanged
                # For example, N4 -> seg = N, el = 04
                seg = seg_part
                el = digits.zfill(2)
            seg = Item.normalize_segment(seg)
            return seg, el, ""
        
        # Nothing matched → return empty fields
        return "", "", ""


class SpreadsheetParser:
    """Parser for Excel spreadsheet files"""
    
    def __init__(self, database, language: str = "UA"):
        """
        Initialize parser
        
        Args:
            database: Database instance for matching items
            language: Language code ("UA" or "EN") for error messages
        """
        self.database = database
        self.language = language
        self.t = TRANSLATIONS.get(language, TRANSLATIONS["UA"])
    
    def parse(self, file_path: Path) -> Tuple[List[Item], bool, Optional[str]]:
        """
        Parse spreadsheet file and create Item objects
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Tuple of (items_list, success, error_message)
        """
        items: List[Item] = []
        all_errors: List[str] = []
        
        try:
            # Load workbook
            workbook = load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            # Get max column (starting from column B = 2)
            max_col = sheet.max_column
            
            # Iterate through columns starting from B (index 2)
            for col_idx in range(2, max_col + 1):
                item = Item()
                column_errors = []
                
                # Get values from rows 1-5
                row1_value = self._get_cell_value(sheet, 1, col_idx)
                row2_value = self._get_cell_value(sheet, 2, col_idx)
                row3_value = self._get_cell_value(sheet, 3, col_idx)
                row4_value = self._get_cell_value(sheet, 4, col_idx)
                row5_value = self._get_cell_value(sheet, 5, col_idx)
                
                # Row 1: spreadsheet_label (required)
                if not row1_value or str(row1_value).strip() == "":
                    column_errors.append(
                        f"{self.t['error_column']} {self._column_letter(col_idx)}: "
                        f"{self.t['error_empty_field'].format(field='spreadsheet_label')}"
                    )
                else:
                    item.spreadsheet_label = str(row1_value).strip()
                
                # Row 2: spreadsheet_edi_info_text (required)
                if not row2_value or str(row2_value).strip() == "":
                    column_errors.append(
                        f"{self.t['error_column']} {self._column_letter(col_idx)}: "
                        f"{self.t['error_empty_field'].format(field='spreadsheet_edi_info_text')}"
                    )
                else:
                    item.spreadsheet_edi_info_text = str(row2_value).strip()
                
                # Row 3: spreadsheet_usage (required)
                if not row3_value or str(row3_value).strip() == "":
                    column_errors.append(
                        f"{self.t['error_column']} {self._column_letter(col_idx)}: "
                        f"{self.t['error_empty_field'].format(field='spreadsheet_usage')}"
                    )
                else:
                    item.spreadsheet_usage = str(row3_value).strip()
                
                # Row 4: spreadsheet_min_max_text (optional)
                if row4_value and str(row4_value).strip():
                    item.spreadsheet_min_max_text = str(row4_value).strip()
                    # Parse min/max only if field is not empty
                    min_max_errors = self._parse_min_max(item, col_idx)
                    column_errors.extend(min_max_errors)
                else:
                    # If field is empty and spreadsheet_label ends with "UOM", set min=2, max=2
                    if item.spreadsheet_label and item.spreadsheet_label.strip().upper().endswith("UOM"):
                        item.spreadsheet_min = 2
                        item.spreadsheet_max = 2
                
                # Row 5: spreadsheet_description (optional)
                if row5_value:
                    item.spreadsheet_description = str(row5_value).strip()
                
                # If there are errors in required fields, skip further processing
                if column_errors:
                    item.parsing_errors = column_errors
                    all_errors.extend(column_errors)
                    items.append(item)
                    continue
                
                # Parse spreadsheet_edi_info_text_cleared
                try:
                    item.spreadsheet_edi_info_text_cleared = Item.clear_edi_info(item.spreadsheet_edi_info_text)
                except Exception as e:
                    error_msg = (
                        f"{self.t['error_column']} {self._column_letter(col_idx)}: "
                        f"{self.t['error_parse_spreadsheet_edi_info']}: {str(e)}"
                    )
                    column_errors.append(error_msg)
                    all_errors.append(error_msg)
                    item.parsing_errors = column_errors
                    items.append(item)
                    continue
                
                # Special handling for N104 with conditions in parentheses
                special_n104_handled = False
                cleared_text = item.spreadsheet_edi_info_text_cleared.strip()

                if cleared_text:
                    # Case 1: N104 (N101=VN and N103=92) -> segment N1, element 04, qualifier taken from N101
                    m = re.match(
                        r'^N104\s*\(\s*N101\s*=\s*([A-Za-z0-9]+)\s+and\s+N103\s*=\s*([A-Za-z0-9]+)\s*\)$',
                        cleared_text,
                        re.IGNORECASE,
                    )
                    if m:
                        qual_from_n101 = m.group(1).strip()
                        item.edi_segment = "N1"
                        item.edi_element_number = "04"
                        item.edi_qualifier = qual_from_n101
                        special_n104_handled = True
                    else:
                        # Case 2: N104 (N103=92) -> segment N1, element 04, qualifier is inherited
                        m = re.match(
                            r'^N104\s*\(\s*N103\s*=\s*([A-Za-z0-9]+)\s*\)$',
                            cleared_text,
                            re.IGNORECASE,
                        )
                        if m:
                            # Take the previous Item if it exists
                            if items and items[-1].edi_segment == "N1":
                                prev_item = items[-1]
                                item.edi_segment = "N1"
                                item.edi_element_number = "04"
                                item.edi_qualifier = prev_item.edi_qualifier
                                special_n104_handled = True
                            else:
                                # If there is no previous N1 Item, treat it as a parsing error
                                error_msg = (
                                    f"{self.t['error_column']} {self._column_letter(col_idx)}: "
                                    f"{self.t['error_n104_missing_previous_qualifier']}"
                                )
                                column_errors.append(error_msg)
                                all_errors.append(error_msg)
                                item.parsing_errors = column_errors
                                items.append(item)
                                continue

                # Parse EDI info (for all other cases)
                if not special_n104_handled:
                    try:
                        edi_segment, edi_element_number, edi_qualifier = Item.parse_edi_info(
                            item.spreadsheet_edi_info_text_cleared
                        )
                        item.edi_segment = edi_segment
                        item.edi_element_number = edi_element_number
                        item.edi_qualifier = edi_qualifier
                    except Exception as e:
                        error_msg = (
                            f"{self.t['error_column']} {self._column_letter(col_idx)}: "
                            f"{self.t['error_parse_edi_info']}: {str(e)}"
                        )
                        column_errors.append(error_msg)
                        all_errors.append(error_msg)
                        item.parsing_errors = column_errors
                        items.append(item)
                        continue
                
                # Match with database
                if item.edi_segment and item.edi_element_number:
                    match_errors = self._match_with_database(item, col_idx)
                    column_errors.extend(match_errors)
                    all_errors.extend(match_errors)
                
                item.parsing_errors = column_errors
                items.append(item)
            
            workbook.close()
            
            # Check if parsing was successful
            success = len(all_errors) == 0
            error_message = "\n".join(all_errors) if all_errors else None
            
            return items, success, error_message
            
        except Exception as e:
            return [], False, f"{self.t['error_read_file']}: {str(e)}"
    
    def _get_cell_value(self, sheet, row: int, col: int) -> Optional[str]:
        """Get cell value as string"""
        cell = sheet.cell(row=row, column=col)
        if cell.value is None:
            return None
        return str(cell.value)
    
    def _column_letter(self, col_idx: int) -> str:
        """Convert column index to letter (1=A, 2=B, etc.)"""
        return openpyxl.utils.get_column_letter(col_idx)
    
    def _parse_min_max(self, item: Item, col_idx: int) -> List[str]:
        """Parse min/max from spreadsheet_min_max_text (optional field)"""
        errors = []
        text = item.spreadsheet_min_max_text.strip()
        
        # Format: "min=1, max=50" or similar
        min_match = re.search(r'min\s*=\s*(\d+)', text, re.IGNORECASE)
        max_match = re.search(r'max\s*=\s*(\d+)', text, re.IGNORECASE)
        
        # Check if text contains min= or max= keywords (field is optional, so only validate if keywords are present)
        has_min_keyword = bool(re.search(r'min\s*=', text, re.IGNORECASE))
        has_max_keyword = bool(re.search(r'max\s*=', text, re.IGNORECASE))
        
        if min_match:
            try:
                item.spreadsheet_min = int(min_match.group(1))
            except ValueError:
                errors.append(
                    f"{self.t['error_column']} {self._column_letter(col_idx)}: "
                    f"{self.t['error_invalid_min_format']} '{text}'"
                )
        elif has_min_keyword:
            # Only show error if min= keyword is present but value is invalid
            errors.append(
                f"{self.t['error_column']} {self._column_letter(col_idx)}: "
                f"{self.t['error_min_not_found']} '{text}'"
            )
        
        if max_match:
            try:
                item.spreadsheet_max = int(max_match.group(1))
            except ValueError:
                errors.append(
                    f"{self.t['error_column']} {self._column_letter(col_idx)}: "
                    f"{self.t['error_invalid_max_format']} '{text}'"
                )
        elif has_max_keyword:
            # Only show error if max= keyword is present but value is invalid
            errors.append(
                f"{self.t['error_column']} {self._column_letter(col_idx)}: "
                f"{self.t['error_max_not_found']} '{text}'"
            )
        
        return errors
    
    def _match_with_database(self, item: Item, col_idx: int) -> List[str]:
        """Match item with database records"""
        errors = []
        
        # Get all items from database
        db_items = self.database.get_all_items()
        
        # Normalize edi_element_number for matching
        edi_element_num = item.edi_element_number
        
        # Special case: if edi_segment=PO1 and edi_element_number>06, treat as 07
        if item.edi_segment == "PO1" and edi_element_num:
            try:
                if int(edi_element_num) > 6:
                    edi_element_num = "07"
            except ValueError:
                # If edi_element_number is not a valid number, keep original value
                pass
        
        # Find matches
        matches = []
        for db_item in db_items:
            db_segment = db_item.get("edi_segment", "")
            db_element_raw = db_item.get("edi_element_number", "")
            db_qualifier = db_item.get("edi_qualifier") or ""
            item_qualifier = item.edi_qualifier or ""
            
            # Normalize element numbers for comparison (convert both to int)
            # Database stores as INTEGER (e.g., 2), parser returns as string with leading zeros (e.g., "02")
            try:
                db_element_int = int(db_element_raw) if db_element_raw != "" else None
                item_element_int = int(edi_element_num) if edi_element_num else None
                element_match = db_element_int == item_element_int if (db_element_int is not None and item_element_int is not None) else False
            except (ValueError, TypeError):
                # If conversion fails, fall back to string comparison
                db_element = str(db_element_raw)
                element_match = db_element == edi_element_num
            
            if db_segment == item.edi_segment and element_match:
                # Match qualifier if both are present, or if both are empty
                if (db_qualifier and item_qualifier and db_qualifier == item_qualifier) or \
                   (not db_qualifier and not item_qualifier):
                    matches.append(db_item)
        
        if len(matches) == 0:
            errors.append(
                f"{self.t['error_column']} {self._column_letter(col_idx)}: "
                f"{self.t['error_no_match']} "
                f"edi_segment={item.edi_segment}, edi_element_number={item.edi_element_number}, "
                f"edi_qualifier={item.edi_qualifier or self.t['error_empty_qualifier']}"
            )
        elif len(matches) > 1:
            errors.append(
                f"{self.t['error_column']} {self._column_letter(col_idx)}: "
                f"{self.t['error_multiple_matches']} "
                f"edi_segment={item.edi_segment}, edi_element_number={item.edi_element_number}, "
                f"edi_qualifier={item.edi_qualifier or self.t['error_empty_qualifier']}"
            )
        else:
            # Single match - fill item properties
            match = matches[0]
            item.item_properties_id = match.get("item_properties_id")
            item.tli_value = match.get("TLI_value", "")
            item.rsx_tag_850 = match.get("850_RSX_tag", "")
            item.tli_tag_850 = match.get("850_TLI_tag", "")
            item.extra_record_defining_rsx_tag = match.get("extra_record_defining_rsx_tag") or ""
            item.extra_record_defining_qual = match.get("extra_record_defining_qual") or ""
            item.is_on_detail_level = bool(match.get("is_on_detail_level", False))
            item.is_partnumber = bool(match.get("is_partnumber", False))
            item.rsx_path_855 = match.get("855_RSX_path", "")
            item.rsx_path_856 = match.get("856_RSX_path", "")
            item.rsx_path_810 = match.get("810_RSX_path", "")
            item.put_in_855 = bool(match.get("put_in_855_by_default", False))
            item.put_in_856 = bool(match.get("put_in_856_by_default", False))
            item.put_in_810 = bool(match.get("put_in_810_by_default", False))

            # Get sourcing group and order path info
            sourcing_group_id = match.get("sourcing_group_properties_id")
            if sourcing_group_id:
                sg_row = self.database.get_sourcing_group(sourcing_group_id)
                if sg_row:
                    order_path_id = sg_row.get("order_path_properties_id")
                    source_path_obj: Optional[SourceFromTLIPath] = None

                    if order_path_id:
                        item.order_path_properties_id = order_path_id
                        op_row = self.database.get_order_path(order_path_id)
                        if op_row:
                            source_path_obj = SourceFromTLIPath(
                                order_path_properties_id=op_row.get("order_path_properties_id"),
                                order_path=op_row.get("order_path", ""),
                                order_change_path=op_row.get("order_change_path", ""),  # legacy
                                java_code_wrapper=op_row.get("java_code_wrapper", ""),  # legacy
                                xtl_part_to_replace_850=op_row.get("xtl_part_to_replace_850", ""),
                                xtl_part_to_paste_850=op_row.get("xtl_part_to_paste_850", ""),
                                xtl_part_to_replace_860=op_row.get("xtl_part_to_replace_860", ""),
                                xtl_part_to_paste_860=op_row.get("xtl_part_to_paste_860", ""),
                            )

                    item.sourcing_group = SourcingGroup(
                        sourcing_group_properties_id=sg_row.get("sourcing_group_properties_id"),
                        populate_method_name=sg_row.get("populate_method_name", ""),
                        map_name=sg_row.get("map_name", ""),
                        order_path_properties_id=sg_row.get("order_path_properties_id"),
                        call_method_java_code=sg_row.get("call_method_java_code", ""),
                        source_from_tli_path=source_path_obj,
                    )
        
        return errors

